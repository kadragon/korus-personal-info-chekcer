"""
이 모듈은 날짜 조작, 디렉토리 생성, 엑셀 파일 처리(열 너비 자동 맞춤 저장),
처리할 특정 엑셀 파일 검색 및 준비와 같은 일반적인 작업을 위한
유틸리티 함수를 제공합니다.
"""

import os
import zipfile
from datetime import datetime

import holidays
import openpyxl
import pandas as pd
from dateutil.relativedelta import relativedelta
from openpyxl.utils import get_column_letter

import config as cfg
from display import (
    print_error,
    print_info,
    print_result,
    print_zip_result,
    print_zip_warning,
)


def get_prev_month_yyyymm() -> str:
    """
    현재 날짜로부터 이전 달을 계산하고 'YYYYMM' 형식의 문자열로 반환합니다.

    반환 값:
        str: 'YYYYMM' 형식의 이전 달입니다.
    """
    today = datetime.today()
    prev_month_date = today - relativedelta(months=1)
    return prev_month_date.strftime("%Y%m")


def make_save_dir(base_save_dir: str) -> str:
    """
    `base_save_dir` 내에 이전 달(YYYYMM)의 이름으로 하위 디렉토리를 생성합니다.
    하위 디렉토리가 이미 존재하면 아무 작업도 수행하지 않습니다.

    매개변수:
        base_save_dir (str): 새 하위 디렉토리가 생성될 기본 디렉토리입니다.
                             이 경로는 기존 디렉토리여야 합니다.

    반환 값:
        str: 생성되었거나 이미 존재하는 이전 달의 하위 디렉토리에 대한 전체 경로입니다.
    """
    prev_month_str = get_prev_month_yyyymm()
    save_dir = os.path.join(base_save_dir, prev_month_str)

    if not os.path.exists(save_dir):
        os.makedirs(save_dir, exist_ok=True)
        # 이 함수는 메인 헤더가 인쇄되기 전에 호출되므로 간단한 인쇄가 더 좋습니다.
        print(f"폴더 생성: {save_dir}")

    return save_dir


def save_excel_with_autofit(df: pd.DataFrame, path: str):
    """
    Pandas DataFrame을 Excel 파일로 저장하고 열 너비를 자동으로 맞춥니다.

    매개변수:
        df (pd.DataFrame): 저장할 DataFrame입니다.
        path (str): Excel 파일이 저장될 전체 경로(파일 이름 포함)입니다.
    """
    df.to_excel(path, index=False)
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    if ws is None:
        wb.close()
        print_error("활성 워크시트를 찾을 수 없어 열 너비를 자동 맞춤할 수 없습니다.")
        return

    for idx, column_cells in enumerate(ws.columns):  # type: ignore
        max_length = 0
        column_letter = get_column_letter(idx + 1)

        for cell in column_cells:
            try:
                if cell.value is not None:
                    cell_value_str = str(cell.value)
                    max_length = max(max_length, len(cell_value_str))
            except Exception as e:
                print_error(f"[열 너비 자동 맞춤] {cell.coordinate}에서 예외 발생: {e}")

        adjusted_width = max_length + 2 if max_length > 0 else 10
        ws.column_dimensions[column_letter].width = adjusted_width  # type: ignore

    wb.save(path)
    wb.close()


def _find_excel_files(download_dir: str, file_prefix: str) -> list[str]:
    """지정된 디렉토리에서 접두사와 확장자를 기준으로 Excel 파일 목록을 찾습니다."""
    if not download_dir or not os.path.isdir(download_dir):
        raise EnvironmentError(f"다운로드 디렉토리를 찾을 수 없습니다: {download_dir}")

    return [
        f
        for f in os.listdir(download_dir)
        if f.startswith(file_prefix) and f.lower().endswith(cfg.EXCEL_EXTENSIONS)
    ]


def _merge_and_preprocess_files(
    excel_files: list[str], download_dir: str
) -> pd.DataFrame | None:
    """Excel 파일 목록을 읽고 단일 데이터프레임으로 병합한 후 전처리합니다."""
    all_dfs = []
    for file_name in excel_files:
        file_path = os.path.join(download_dir, file_name)
        try:
            if file_path.lower().endswith(".xlsx"):
                df = pd.read_excel(file_path)
            else:
                df = pd.read_excel(file_path, engine="xlrd")
            all_dfs.append(df)
        except Exception as e:
            print_error(f"'{file_name}' 파일 처리 중 오류 발생: {e}")
            return None

    if not all_dfs:
        return None

    merged_df = pd.concat(all_dfs, ignore_index=True)

    # "접속일시" 컬럼이 존재하면 datetime으로 변환
    if cfg.COL_ACCESS_TIME in merged_df.columns:
        merged_df[cfg.COL_ACCESS_TIME] = pd.to_datetime(merged_df[cfg.COL_ACCESS_TIME])

    # "교번" 또는 "신분번호" 컬럼을 "교직원ID"로 표준화
    if "교번" in merged_df.columns:
        merged_df.rename(columns={"교번": cfg.COL_EMPLOYEE_ID}, inplace=True)
    elif "신분번호" in merged_df.columns:
        merged_df.rename(columns={"신분번호": cfg.COL_EMPLOYEE_ID}, inplace=True)

    return merged_df


def find_and_prepare_excel_file(
    download_dir: str,
    file_prefix: str,
    save_dir: str,
    output_file_basename: str,
    prev_month: str,
) -> tuple[pd.DataFrame | None, str | None]:
    """
    지정된 폴더에서 Excel 파일을 찾아 병합, 전처리 및 저장합니다.

    이 함수는 다음을 수행합니다:
    1. `_find_excel_files`를 사용하여 관련 파일들을 찾습니다.
    2. `_merge_and_preprocess_files`를 사용하여 파일들을 병합하고 전처리합니다.
    3. 병합된 데이터프레임을 중간 결과물로 저장합니다.
    """
    try:
        excel_files = _find_excel_files(download_dir, file_prefix)
        if not excel_files:
            print_info(
                f"'{file_prefix}'로 시작하는 파일을 찾을 수 없습니다. "
                f"이 검사는 건너뜁니다."
            )
            return None, None
    except EnvironmentError as e:
        print_error(str(e))
        return None, None

    merged_df = _merge_and_preprocess_files(excel_files, download_dir)
    if merged_df is None:
        return None, None

    print_info(f"{output_file_basename} 원본 데이터: {len(merged_df)}건")

    os.makedirs(save_dir, exist_ok=True)
    destination_save_path = os.path.join(
        save_dir, f"{output_file_basename}_{prev_month}.xlsx"
    )
    try:
        merged_df.to_excel(destination_save_path, index=False)
        save_msg = (
            f"모든 파일을 합쳐 '{os.path.basename(destination_save_path)}'"
            f"(으)로 저장했습니다."
        )
        print_info(save_msg)
    except Exception as e:
        print_error(f"병합된 파일 저장 중 오류 발생: {e}")
        return None, None

    return merged_df, destination_save_path


def zip_files_by_prefix(target_dir: str, prefix_list: list[str]):
    """
    파일명에서 '_' 앞부분(붙임N ... )을 zip 이름으로 하여 압축 생성
    """
    files = [f for f in os.listdir(target_dir) if f.endswith(".xlsx")]

    for prefix in prefix_list:
        matched = [f for f in files if f.startswith(prefix)]
        if not matched:
            print_zip_warning(prefix)
            continue

        group_name = matched[0].split("_")[0].split("(")[0]
        zip_name = f"{group_name}.zip"
        zip_path = os.path.join(target_dir, zip_name)

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            for filename in matched:
                zipf.write(os.path.join(target_dir, filename), arcname=filename)

        print_zip_result(zip_name, len(matched))


def filter_by_time_conditions(
    df: pd.DataFrame,
    time_col: str,
    employee_id_col: str,
    check_off_hours: bool,
    check_holidays_weekends: bool,
    off_hours_start: int,
    off_hours_end: int,
) -> pd.DataFrame:
    """
    시간 조건(업무 시간 외, 공휴일/주말)에 따라 데이터프레임을 필터링합니다.

    매개변수:
        df (pd.DataFrame): 필터링할 데이터프레임.
        time_col (str): 타임스탬프 정보를 포함하는 컬럼 이름.
        employee_id_col (str): 직원 ID를 포함하는 컬럼 이름.
        check_off_hours (bool): 업무 시간 외 검사를 활성화할지 여부.
        check_holidays_weekends (bool): 공휴일 및 주말 검사를 활성화할지 여부.
        off_hours_start (int): 업무 시간 외 시작 시간.
        off_hours_end (int): 업무 시간 외 종료 시간.

    반환 값:
        pd.DataFrame: 지정된 시간 조건을 충족하는 필터링된 데이터프레임.
    """
    if df is None:
        raise ValueError("Input DataFrame cannot be None.")

    df_copy = df.copy()

    final_mask = pd.Series(False, index=df.index)

    if check_off_hours:
        hour = df_copy[time_col].dt.hour
        is_off_hour = (hour < off_hours_end) | (hour >= off_hours_start)
        final_mask |= is_off_hour

    if check_holidays_weekends:
        years = df_copy[time_col].dt.year.unique()
        kr_holidays = holidays.KR(years=years)  # type: ignore [attr-defined]
        weekday = df_copy[time_col].dt.weekday
        date_only = df_copy[time_col].dt.date

        is_weekend = weekday >= 5  # Monday is 0, Sunday is 6
        is_holiday = date_only.isin(kr_holidays)
        final_mask |= is_weekend
        final_mask |= is_holiday

    return df_copy[final_mask].sort_values([employee_id_col, time_col])


def run_and_save_check(
    df: pd.DataFrame,
    check_func,
    save_path: str,
    result_description: str,
):
    """
    검사 함수를 실행하고, 결과가 있으면 Excel 파일로 저장한 후 상태 메시지를 출력합니다.

    매개변수:
        df (pd.DataFrame): 검사를 수행할 입력 DataFrame입니다.
        check_func (function): (
            DataFrame을 인자로 받아 필터링된 DataFrame을 반환하는 함수입니다.
        )
        save_path (str): 결과 Excel 파일이 저장될 경로입니다.
        result_description (str): (
            결과가 발견되었거나 발견되지 않았을 때 출력할 메시지에 사용될 설명입니다.
        )
    """
    filtered_df = check_func(df)
    if not filtered_df.empty:
        save_excel_with_autofit(filtered_df, save_path)
        print_result(
            is_detected=True,
            description=result_description,
            filename=os.path.basename(save_path),
        )
    else:
        print_result(is_detected=False, description=result_description)
